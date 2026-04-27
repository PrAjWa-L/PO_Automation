"""
Payments & UTR API
  GET    /api/payments                       — list all (filters: status, po_id)
  POST   /api/payments                       — record a payment
  GET    /api/payments/<id>                  — get one
  PUT    /api/payments/<id>                  — update (e.g. add UTR after initiation)
  PATCH  /api/payments/<id>/utr              — shortcut: just record UTR + mark Paid
  DELETE /api/payments/<id>                  — delete only Pending payments
  GET    /api/payments/pending-utr           — list payments missing UTR
  GET    /api/payments/summary               — totals by status
"""

from flask import Blueprint, request
from app import db
from app.models import Payment, PurchaseOrder
from app.utils import ok, created, err, not_found, server_err, parse_date
from app.auth import login_required, role_required, current_user
from datetime import datetime, timezone, date
import io
from flask import send_file

payments_bp = Blueprint("payments", __name__)

VALID_TYPES  = ["Advance", "Partial", "Full", "Final"]
VALID_MODES  = ["NEFT", "RTGS", "IMPS", "Cheque", "Cash", "UPI"]
VALID_STATUS = ["Pending", "Paid", "Failed", "Cancelled"]


# ─── LIST (grouped by PO) ────────────────────────────────────
@payments_bp.get("")
def list_payments():
    from sqlalchemy import func
    from collections import OrderedDict

    status = request.args.get("status")
    po_id  = request.args.get("po_id")
    mode   = request.args.get("payment_mode")

    q = Payment.query
    if status: q = q.filter(Payment.status == status)
    if po_id:  q = q.filter(Payment.po_id  == po_id)
    if mode:   q = q.filter(Payment.payment_mode == mode)

    payments = q.order_by(Payment.po_id, Payment.payment_date.asc()).all()

    # Total paid per PO (Paid status only)
    paid_by_po = dict(
        db.session.query(Payment.po_id, func.sum(Payment.amount))
        .filter(Payment.status == "Paid")
        .group_by(Payment.po_id)
        .all()
    )

    # Group payments by PO
    groups = OrderedDict()
    for p in payments:
        pid = p.po_id
        if pid not in groups:
            po          = p.po
            grand_total = float(po.grand_total or 0) if po else 0
            total_paid  = float(paid_by_po.get(pid, 0) or 0)
            groups[pid] = {
                "po_id":       pid,
                "vendor_name": po.vendor_name if po else None,
                "department":  po.department  if po else None,
                "grand_total": grand_total,
                "total_paid":  round(total_paid, 2),
                "balance":     round(max(0, grand_total - total_paid), 2),
                "payments":    [],
            }
        groups[pid]["payments"].append(p.to_dict())

    # Sort by most-recent payment date first
    def _latest(g):
        dates = [p["payment_date"] for p in g["payments"] if p["payment_date"]]
        return max(dates) if dates else ""

    result = sorted(groups.values(), key=_latest, reverse=True)
    return ok(result)


# ─── PENDING UTR (special list — no UTR + status Pending) ────
@payments_bp.get("/pending-utr")
def pending_utr():
    payments = (
        Payment.query
        .filter(Payment.status == "Pending")
        .filter(
            db.or_(Payment.utr_number == None, Payment.utr_number == "")
        )
        .order_by(Payment.due_date.asc().nullslast())
        .all()
    )
    return ok([p.to_dict() for p in payments])


# ─── SUMMARY ─────────────────────────────────────────────────
@payments_bp.get("/summary")
def summary():
    from sqlalchemy import func
    rows = (
        db.session.query(Payment.status, func.sum(Payment.amount))
        .group_by(Payment.status)
        .all()
    )
    result = {status: float(total or 0) for status, total in rows}
    return ok(result)


# ─── GET ONE ─────────────────────────────────────────────────
@payments_bp.get("/<int:pid>")
def get_payment(pid):
    p = Payment.query.get(pid)
    if not p:
        return not_found("Payment")
    return ok(p.to_dict())


# ─── CREATE ──────────────────────────────────────────────────
@payments_bp.post("")
def create_payment():
    data = request.get_json(silent=True) or {}

    po_id = (data.get("po_id") or "").strip()
    if not po_id:
        return err("po_id is required")

    po = PurchaseOrder.query.get(po_id)
    if not po:
        return not_found("Purchase Order")
    if po.status not in ("Approved", "Closed"):
        return err(
            f"Payments can only be recorded against Approved or Closed POs "
            f"(current status: {po.status})", 409
        )

    try:
        amount = float(data.get("amount", 0) or 0)
    except (TypeError, ValueError):
        return err("amount must be a number")

    if amount <= 0:
        return err("amount must be greater than 0")

    # Guard: don't allow overpayment beyond grand total
    existing_paid = sum(
        float(p.amount or 0)
        for p in po.payments.filter(Payment.status == "Paid").all()
    )
    if existing_paid + amount > float(po.grand_total or 0) + 0.01:   # 1 paisa tolerance
        return err(
            f"Payment would exceed PO grand total. "
            f"Grand total: ₹{po.grand_total}, Already paid: ₹{round(existing_paid,2)}, "
            f"Balance: ₹{round(float(po.grand_total or 0) - existing_paid, 2)}", 409
        )

    pdate = parse_date(data.get("payment_date"))
    if not pdate:
        return err("payment_date is required (YYYY-MM-DD)")

    payment_type = data.get("payment_type", "Full")
    if payment_type not in VALID_TYPES:
        return err(f"payment_type must be one of: {', '.join(VALID_TYPES)}")

    payment_mode = (data.get("payment_mode") or "").strip()
    if payment_mode and payment_mode not in VALID_MODES:
        return err(f"payment_mode must be one of: {', '.join(VALID_MODES)}")

    try:
        p = Payment(
            po_id        = po_id,
            payment_type = payment_type,
            amount       = amount,
            payment_date = pdate,
            due_date     = parse_date(data.get("due_date")),
            utr_number   = (data.get("utr_number") or "").strip() or None,
            payment_mode = payment_mode or None,
            bank_ref     = (data.get("bank_ref")  or "").strip() or None,
            cheque_no    = (data.get("cheque_no") or "").strip() or None,
            status          = "Pending",
            approval_status = "Pending Approval",
            remarks         = (data.get("remarks") or "").strip() or None,
        )
        db.session.add(p)
        db.session.commit()
        return created(p.to_dict(), f"Payment recorded for {po_id}")
    except Exception as e:
        db.session.rollback()
        return server_err(e)


# ─── UPDATE (general) ────────────────────────────────────────
@payments_bp.put("/<int:pid>")
def update_payment(pid):
    p = Payment.query.get(pid)
    if not p:
        return not_found("Payment")

    if p.status == "Paid":
        return err("Cannot edit a Paid payment. Use /utr endpoint to add UTR only.", 409)

    data = request.get_json(silent=True) or {}

    try:
        if "amount"       in data: p.amount       = float(data["amount"] or 0)
        if "payment_date" in data: p.payment_date = parse_date(data["payment_date"]) or p.payment_date
        if "due_date"     in data: p.due_date      = parse_date(data["due_date"])
        if "payment_type" in data:
            if data["payment_type"] not in VALID_TYPES:
                return err(f"payment_type must be one of: {', '.join(VALID_TYPES)}")
            p.payment_type = data["payment_type"]
        if "payment_mode" in data:
            if data["payment_mode"] and data["payment_mode"] not in VALID_MODES:
                return err(f"payment_mode must be one of: {', '.join(VALID_MODES)}")
            p.payment_mode = data["payment_mode"]
        if "utr_number"   in data: p.utr_number   = (data["utr_number"]  or "").strip() or None
        if "bank_ref"     in data: p.bank_ref      = (data["bank_ref"]    or "").strip() or None
        if "cheque_no"    in data: p.cheque_no     = (data["cheque_no"]   or "").strip() or None
        if "status"       in data:
            if data["status"] not in VALID_STATUS:
                return err(f"status must be one of: {', '.join(VALID_STATUS)}")
            p.status = data["status"]
        if "remarks"      in data: p.remarks       = data["remarks"]

        db.session.commit()
        return ok(p.to_dict(), "Payment updated")
    except Exception as e:
        db.session.rollback()
        return server_err(e)


# ─── UTR SHORTCUT ────────────────────────────────────────────
@payments_bp.patch("/<int:pid>/utr")
def record_utr(pid):
    """
    Quickest workflow: bank sends UTR → accounts team records it here
    → payment auto-transitions to Paid.
    """
    p = Payment.query.get(pid)
    if not p:
        return not_found("Payment")

    if p.status == "Paid":
        return err("Payment is already marked Paid", 409)

    if p.approval_status != "Approved":
        return err("Payment must be approved by COO before entering UTR", 403)

    data = request.get_json(silent=True) or {}
    utr  = (data.get("utr_number") or "").strip()
    if not utr:
        return err("utr_number is required")

    try:
        p.utr_number   = utr
        p.bank_ref     = (data.get("bank_ref")   or p.bank_ref  or "").strip() or None
        p.payment_mode = (data.get("payment_mode") or p.payment_mode or "").strip() or None
        p.status       = "Paid"
        if data.get("remarks"):
            p.remarks = data["remarks"]
        db.session.commit()
        return ok(p.to_dict(), f"UTR {utr} recorded — payment marked Paid")
    except Exception as e:
        db.session.rollback()
        return server_err(e)



# ─── COO APPROVE ─────────────────────────────────────────────
@payments_bp.patch("/<int:pid>/approve")
@role_required("coo")
def approve_payment(pid):
    p = Payment.query.get(pid)
    if not p:
        return not_found("Payment")

    if p.approval_status == "Approved":
        return err("Payment is already approved", 409)

    if p.status == "Paid":
        return err("Payment is already paid", 409)

    user = current_user()
    try:
        p.approval_status = "Approved"
        p.approved_by     = user.get("display_name") or user.get("username", "")
        p.approved_at     = datetime.now(timezone.utc)
        db.session.commit()
        return ok(p.to_dict(), f"Payment {pid} approved")
    except Exception as e:
        db.session.rollback()
        return server_err(e)


# ─── COO REJECT ──────────────────────────────────────────────
@payments_bp.patch("/<int:pid>/reject-payment")
@role_required("coo")
def reject_payment(pid):
    p = Payment.query.get(pid)
    if not p:
        return not_found("Payment")

    if p.status == "Paid":
        return err("Cannot reject a paid payment", 409)

    data = request.get_json(silent=True) or {}
    try:
        p.approval_status = "Rejected"
        p.approved_by     = (current_user() or {}).get("display_name", "")
        p.approved_at     = datetime.now(timezone.utc)
        if data.get("remarks"):
            p.remarks = data["remarks"]
        db.session.commit()
        return ok(p.to_dict(), f"Payment {pid} rejected")
    except Exception as e:
        db.session.rollback()
        return server_err(e)


# ─── DELETE ──────────────────────────────────────────────────
@payments_bp.delete("/<int:pid>")
def delete_payment(pid):
    p = Payment.query.get(pid)
    if not p:
        return not_found("Payment")

    if p.status not in ("Pending", "Failed", "Cancelled"):
        return err(
            f"Only Pending/Failed/Cancelled payments can be deleted "
            f"(current: {p.status})", 409
        )
    try:
        db.session.delete(p)
        db.session.commit()
        return ok(msg=f"Payment {pid} deleted")
    except Exception as e:
        db.session.rollback()
        return server_err(e)
    
# ─── PAYMENT REPORT EXPORT ───────────────────────────────────
@payments_bp.get("/export")
@login_required
def export_payments():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return err("openpyxl not installed.", 500)

    from app.models import PurchaseOrder

    payments = (
        db.session.query(Payment)
        .join(PurchaseOrder, Payment.po_id == PurchaseOrder.id)
        .order_by(Payment.payment_date.desc())
        .all()
    )

    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Payment Report"

    # ── Styles ──────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1a56db")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    alt_fill   = PatternFill("solid", fgColor="EEF2FF")
    paid_fill  = PatternFill("solid", fgColor="D1FAE5")   # green tint
    pend_fill  = PatternFill("solid", fgColor="FEF3C7")   # amber tint
    fail_fill  = PatternFill("solid", fgColor="FEE2E2")   # red tint
    bd         = Side(style="thin", color="D1D5DB")
    border     = Border(left=bd, right=bd, top=bd, bottom=bd)

    status_fill = {
        "Paid":      paid_fill,
        "Pending":   pend_fill,
        "Failed":    fail_fill,
        "Cancelled": alt_fill,
    }

    headers = [
        "Payment ID", "PO Number", "Vendor Name", "Department",
        "Payment Type", "Amount (₹)", "Payment Date", "Due Date",
        "Mode", "UTR Number", "Bank Ref", "Cheque No",
        "Status", "Approval Status", "Approved By", "Approved At",
        "PO Grand Total (₹)", "Total Paid (₹)", "Outstanding (₹)", "Remarks",
    ]

    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.font       = hdr_font
        cell.fill       = hdr_fill
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = border
    ws.row_dimensions[1].height = 30

    # Pre-compute total paid per PO
    from sqlalchemy import func
    paid_by_po = dict(
        db.session.query(Payment.po_id, func.sum(Payment.amount))
        .filter(Payment.status == "Paid")
        .group_by(Payment.po_id)
        .all()
    )

    for row_idx, p in enumerate(payments, 2):
        po           = p.po
        total_paid   = float(paid_by_po.get(p.po_id, 0) or 0)
        grand_total  = float(po.grand_total or 0) if po else 0
        outstanding  = round(grand_total - total_paid, 2)
        row_fill     = status_fill.get(p.status, alt_fill if row_idx % 2 == 0 else None)

        row_data = [
            p.id,
            p.po_id,
            po.vendor_name   if po else "—",
            po.department    if po else "—",
            p.payment_type,
            float(p.amount or 0),
            p.payment_date.strftime("%d-%m-%Y")  if p.payment_date else "",
            p.due_date.strftime("%d-%m-%Y")      if p.due_date     else "",
            p.payment_mode   or "",
            p.utr_number     or "",
            p.bank_ref       or "",
            p.cheque_no      or "",
            p.status,
            p.approval_status or "",
            p.approved_by     or "",
            p.approved_at.strftime("%d-%m-%Y %H:%M") if p.approved_at else "",
            grand_total,
            total_paid,
            outstanding,
            p.remarks or "",
        ]

        for col, val in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill

    # Column widths
    col_widths = [12, 16, 28, 14, 14, 16, 14, 12,
                  10, 24, 20, 14, 12, 18, 20, 20,
                  18, 14, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # ── Vendor Summary Sheet ─────────────────────────────────
    ws2 = wb.create_sheet("Vendor Outstanding")

    vs_headers = ["Vendor Name", "Total PO Value (₹)", "Total Paid (₹)", "Outstanding (₹)", "No. of POs"]
    for col, h in enumerate(vs_headers, 1):
        cell           = ws2.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws2.row_dimensions[1].height = 25

    from sqlalchemy import func as f2
    vendor_summary = (
        db.session.query(
            PurchaseOrder.vendor_name,
            f2.count(PurchaseOrder.id),
            f2.sum(PurchaseOrder.grand_total),
        )
        .filter(PurchaseOrder.status == "Approved")
        .group_by(PurchaseOrder.vendor_name)
        .order_by(f2.sum(PurchaseOrder.grand_total).desc())
        .all()
    )

    for row_idx, (vname, po_count, po_total) in enumerate(vendor_summary, 2):
        total_po   = float(po_total or 0)
        total_pd   = float(paid_by_po.get(vname, 0) or 0)

        # Sum paid across all POs for this vendor
        vendor_paid = db.session.query(f2.sum(Payment.amount))\
            .join(PurchaseOrder, Payment.po_id == PurchaseOrder.id)\
            .filter(PurchaseOrder.vendor_name == vname, Payment.status == "Paid")\
            .scalar() or 0
        vendor_paid  = float(vendor_paid)
        outstanding  = round(total_po - vendor_paid, 2)
        row_fill     = paid_fill if outstanding <= 0 else (pend_fill if outstanding < total_po else fail_fill)

        row_data = [vname, total_po, vendor_paid, outstanding, po_count]
        for col, val in enumerate(row_data, 1):
            cell           = ws2.cell(row=row_idx, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            cell.fill      = row_fill

    for i, w in enumerate([30, 20, 18, 18, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Stream ───────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ProcureIQ_Payment_Report_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )