"""
Purchase Orders API
  GET    /api/purchase-orders               — list (filters: status, dept, vendor_id)
  POST   /api/purchase-orders               — create with line items
  GET    /api/purchase-orders/<id>          — get one (includes line items + payments)
  PUT    /api/purchase-orders/<id>          — update header + replace line items
  DELETE /api/purchase-orders/<id>          — delete (only Draft allowed)
  PATCH  /api/purchase-orders/<id>/status   — advance status workflow
  GET    /api/purchase-orders/stats         — dashboard summary numbers
"""
from app.auth import login_required, role_required, current_user
from flask import Blueprint, request, send_file
from app import db
import io
from flask import send_file
from datetime import date
from app.models import PurchaseOrder, LineItem, Vendor
from app.utils import (
    ok, created, err, not_found, server_err,
    next_po_id, parse_date,
)

po_bp = Blueprint("purchase_orders", __name__)

VALID_STATUSES = ["Draft", "Pending Approval", "Approved", "Rejected", "Closed"]
VALID_DEPTS    = ["IT", "Maintenance", "Housekeeping", "Accounts",
                  "HR", "Pharmacy", "Administration"]


# ─── LIST ────────────────────────────────────────────────────
@po_bp.get("")
def list_pos():
    from sqlalchemy import func
    q = PurchaseOrder.query

    status = request.args.get("status")
    dept   = request.args.get("dept")
    vid    = request.args.get("vendor_id")

    if status:
        q = q.filter(PurchaseOrder.status == status)
    if dept:
        q = q.filter(PurchaseOrder.department == dept)
    if vid:
        q = q.filter(PurchaseOrder.vendor_id == vid)

    pos = q.order_by(PurchaseOrder.created_at.desc()).all()

    # Pre-aggregate all counts/totals in bulk — never touch lazy relationships in the loop
    from app.models import Payment, Quotation
    paid_by_po = dict(
        db.session.query(Payment.po_id, func.sum(Payment.amount))
        .filter(Payment.status == "Paid")
        .group_by(Payment.po_id)
        .all()
    )
    pay_count_by_po = dict(
        db.session.query(Payment.po_id, func.count(Payment.id))
        .group_by(Payment.po_id)
        .all()
    )
    quot_count_by_po = dict(
        db.session.query(Quotation.po_id, func.count(Quotation.id))
        .filter(Quotation.po_id != None)
        .group_by(Quotation.po_id)
        .all()
    )

    result = []
    for p in pos:
        d    = p.to_dict(include_items=False)
        paid = float(paid_by_po.get(p.id, 0) or 0)
        gt   = float(p.grand_total or 0)
        d["payments_summary"] = {
            "count":      int(pay_count_by_po.get(p.id, 0) or 0),
            "paid_total": round(paid, 2),
            "balance":    round(max(0, gt - paid), 2),
        }
        d["payments_count"]   = int(pay_count_by_po.get(p.id,   0) or 0)
        d["quotations_count"] = int(quot_count_by_po.get(p.id, 0) or 0)
        result.append(d)

    return ok(result)


# ─── STATS (must come before <id> route) ─────────────────────
@po_bp.get("/stats")
def stats():
    from sqlalchemy import func
    total_pos   = PurchaseOrder.query.count()
    draft       = PurchaseOrder.query.filter_by(status="Draft").count()
    pending     = PurchaseOrder.query.filter_by(status="Pending Approval").count()
    approved    = PurchaseOrder.query.filter_by(status="Approved").count()
    ytd_spend   = db.session.query(
        func.sum(PurchaseOrder.grand_total)
    ).filter(PurchaseOrder.status == "Approved").scalar() or 0

    return ok({
        "total_pos":   total_pos,
        "draft":       draft,
        "pending":     pending,
        "approved":    approved,
        "ytd_spend":   float(ytd_spend),
    })

# ─── DEPT SPEND (for dashboard chart) ────────────────────────
@po_bp.get("/dept-spend")
def dept_spend():
    from sqlalchemy import func
    rows = (
        db.session.query(
            PurchaseOrder.department,
            func.sum(PurchaseOrder.grand_total)
        )
        .filter(PurchaseOrder.status == "Approved")
        .group_by(PurchaseOrder.department)
        .order_by(func.sum(PurchaseOrder.grand_total).desc())
        .all()
    )
    return ok([{"department": dept, "total": float(total or 0)} for dept, total in rows])

# ─── EXPORT APPROVED POs ─────────────────────────────────────
@po_bp.get("/export")
@login_required
def export_approved():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return err("openpyxl is not installed. Run: pip install openpyxl", 500)

    pos = (
        PurchaseOrder.query
        .filter(PurchaseOrder.status == "Approved")
        .order_by(PurchaseOrder.po_date.desc())
        .all()
    )

    wb = openpyxl.Workbook()

    # ── Sheet 1: PO Summary ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "PO Summary"

    header_fill = PatternFill("solid", fgColor="1a56db")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill    = PatternFill("solid", fgColor="EEF2FF")
    border_side = Side(style="thin", color="D1D5DB")
    cell_border = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side
    )

    headers = [
        "PO Number", "PO Date", "Order Type", "Department",
        "Vendor Name", "Vendor GST", "Requested By", "Approved By",
        "Payment Terms", "Subtotal (₹)", "Discount (₹)", "GST (₹)",
        "TDS (₹)", "Grand Total (₹)", "Advance %", "Advance Amt (₹)",
        "Paid (₹)", "Balance (₹)", "Notes"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = cell_border

    ws1.row_dimensions[1].height = 30

    for row_idx, po in enumerate(pos, 2):
        payments     = po.payments.all()
        paid_total   = sum(float(p.amount or 0) for p in payments if p.status == "Paid")
        balance      = float(po.grand_total or 0) - paid_total
        fill         = alt_fill if row_idx % 2 == 0 else None

        row_data = [
            po.id,
            po.po_date.strftime("%d-%m-%Y") if po.po_date else "",
            po.order_type or "Purchase Order",
            po.department,
            po.vendor_name or "",
            po.vendor_gst  or "",
            po.requested_by or "",
            po.approved_by  or "",
            po.payment_terms or "",
            float(po.subtotal    or 0),
            float(po.discount    or 0),
            float(po.gst_total   or 0),
            float(po.tds_amt     or 0),
            float(po.grand_total or 0),
            float(po.advance_pct or 0),
            float(po.advance_amt or 0),
            round(paid_total, 2),
            round(balance,    2),
            po.notes or "",
        ]

        for col, val in enumerate(row_data, 1):
            cell           = ws1.cell(row=row_idx, column=col, value=val)
            cell.border    = cell_border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    # Auto-width
    col_widths = [14, 12, 16, 14, 28, 18, 18, 18, 14,
                  14, 13, 12, 10, 16, 11, 16, 12, 12, 30]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Line Items ──────────────────────────────────
    ws2 = wb.create_sheet("Line Items")

    li_headers = [
        "PO Number", "Vendor Name", "Department", "Item Name",
        "Description", "HSN Code", "Qty", "Unit Price (₹)",
        "Discount %", "GST %", "CGST (₹)", "SGST (₹)", "Line Total (₹)"
    ]
    for col, h in enumerate(li_headers, 1):
        cell           = ws2.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border

    ws2.row_dimensions[1].height = 30
    li_row = 2

    for po in pos:
        for li in po.line_items:
            alt = PatternFill("solid", fgColor="EEF2FF") if li_row % 2 == 0 else None
            row_data = [
                po.id,
                po.vendor_name or "",
                po.department,
                li.item_name,
                li.description or "",
                li.hsn_code    or "",
                float(li.qty          or 1),
                float(li.unit_price   or 0),
                float(li.discount_pct or 0),
                float(li.gst_pct      or 18),
                float(li.cgst         or 0),
                float(li.sgst         or 0),
                float(li.line_total   or 0),
            ]
            for col, val in enumerate(row_data, 1):
                cell           = ws2.cell(row=li_row, column=col, value=val)
                cell.border    = cell_border
                cell.alignment = Alignment(vertical="center")
                if alt:
                    cell.fill = alt
            li_row += 1

    li_widths = [14, 28, 14, 30, 35, 12, 8, 15, 12, 8, 12, 12, 14]
    for i, w in enumerate(li_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Payments ────────────────────────────────────
    ws3 = wb.create_sheet("Payments")

    pay_headers = [
        "PO Number", "Vendor Name", "Payment Type", "Amount (₹)",
        "Payment Date", "Due Date", "Mode", "UTR Number",
        "Status", "Approval Status", "Approved By", "Remarks"
    ]
    for col, h in enumerate(pay_headers, 1):
        cell           = ws3.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border

    ws3.row_dimensions[1].height = 30
    pay_row = 2

    for po in pos:
        for p in po.payments.all():
            alt = PatternFill("solid", fgColor="EEF2FF") if pay_row % 2 == 0 else None
            row_data = [
                po.id,
                po.vendor_name or "",
                p.payment_type,
                float(p.amount or 0),
                p.payment_date.strftime("%d-%m-%Y") if p.payment_date else "",
                p.due_date.strftime("%d-%m-%Y")     if p.due_date     else "",
                p.payment_mode  or "",
                p.utr_number    or "",
                p.status,
                p.approval_status or "",
                p.approved_by     or "",
                p.remarks         or "",
            ]
            for col, val in enumerate(row_data, 1):
                cell           = ws3.cell(row=pay_row, column=col, value=val)
                cell.border    = cell_border
                cell.alignment = Alignment(vertical="center")
                if alt:
                    cell.fill = alt
            pay_row += 1

    pay_widths = [14, 28, 14, 14, 14, 12, 10, 24, 10, 16, 20, 30]
    for i, w in enumerate(pay_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ── Stream to response ───────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ProcureIQ_Approved_POs_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

# ─── GET ONE ─────────────────────────────────────────────────
@po_bp.get("/<string:po_id>")
def get_po(po_id):
    po = PurchaseOrder.query.get(po_id)
    if not po:
        return not_found("Purchase Order")
    d = po.to_dict()
    # Attach payment summary
    payments = po.payments.all()
    paid_total = sum(float(p.amount or 0) for p in payments if p.status == "Paid")
    d["payments_summary"] = {
        "count":      len(payments),
        "paid_total": paid_total,
        "balance":    float(po.grand_total or 0) - paid_total,
    }
    d["payments"] = [p.to_dict() for p in payments]
    return ok(d)


# ─── CREATE ──────────────────────────────────────────────────
@po_bp.post("")
def create_po():
    data = request.get_json(silent=True) or {}

    dept = data.get("department", "").strip()
    if not dept:
        return err("Department is required")
    if dept not in VALID_DEPTS:
        return err(f"Invalid department. Choose from: {', '.join(VALID_DEPTS)}")

    po_date = parse_date(data.get("po_date"))
    if not po_date:
        return err("po_date is required (YYYY-MM-DD)")

    items_data = data.get("line_items", [])
    if not items_data:
        return err("At least one line item is required")

    item_errors = _validate_items(items_data)
    if item_errors:
        return err("Line item validation failed", details=item_errors)

    try:
        # ── Auto-fill vendor snapshot ──────────────────────
        vendor_id   = data.get("vendor_id")
        vendor_name = data.get("vendor_name", "").strip()
        vendor_gst  = data.get("vendor_gst",  "").strip()
        vendor_addr = data.get("vendor_addr", "").strip()
        vendor_bank = data.get("vendor_bank", "").strip()

        if vendor_id:
            v = db.session.get(Vendor, vendor_id)
            if v:
                vendor_name = vendor_name or v.name
                vendor_gst  = vendor_gst  or (v.gst  or "")
                vendor_addr = vendor_addr or (v.address or "")
                vendor_bank = vendor_bank or _fmt_bank(v)

        po = PurchaseOrder(
            id            = next_po_id(),
            vendor_id     = vendor_id,
            vendor_name   = vendor_name,
            vendor_gst    = vendor_gst,
            vendor_addr   = vendor_addr,
            vendor_bank   = vendor_bank,
            department    = dept,
            requested_by  = data.get("requested_by", "").strip(),
            created_by    = data.get("created_by",   "Accounts Team").strip(),
            approved_by   = data.get("approved_by",  "").strip(),
            po_date       = po_date,
            delivery_date = parse_date(data.get("delivery_date")),
            payment_terms = data.get("payment_terms", "Net 30"),
            notes         = data.get("notes", "").strip(),
            status        = data.get("status", "Draft"),
            advance_pct   = float(data.get("advance_pct", 0) or 0),
        )
        db.session.add(po)
        db.session.flush()   # get po.id before adding children

        for i, item in enumerate(items_data):
            li = _build_line_item(po.id, item, i)
            db.session.add(li)

        db.session.flush()
        po.recalculate_totals()
        db.session.commit()
        return created(po.to_dict(), f"Purchase Order {po.id} created")

    except Exception as e:
        db.session.rollback()
        return server_err(e)


# ─── UPDATE ──────────────────────────────────────────────────
@po_bp.put("/<string:po_id>")
def update_po(po_id):
    po = PurchaseOrder.query.get(po_id)
    if not po:
        return not_found("Purchase Order")

    if po.status == "Closed":
        return err("Cannot edit a Closed PO", 409)

    data = request.get_json(silent=True) or {}

    items_data = data.get("line_items")
    if items_data is not None:
        if not items_data:
            return err("At least one line item is required")
        item_errors = _validate_items(items_data)
        if item_errors:
            return err("Line item validation failed", details=item_errors)

    try:
        # Update header fields if provided
        if "department"    in data: po.department    = data["department"]
        if "requested_by"  in data: po.requested_by  = data["requested_by"]
        if "created_by"    in data: po.created_by    = data["created_by"]
        if "approved_by"   in data: po.approved_by   = data["approved_by"]
        if "po_date"       in data: po.po_date        = parse_date(data["po_date"]) or po.po_date
        if "delivery_date" in data: po.delivery_date  = parse_date(data["delivery_date"])
        if "payment_terms" in data: po.payment_terms  = data["payment_terms"]
        if "notes"         in data: po.notes          = data["notes"]
        if "advance_pct"   in data: po.advance_pct    = float(data["advance_pct"] or 0)
        if "vendor_id"     in data:
            po.vendor_id = data["vendor_id"]
            v = Vendor.query.get(po.vendor_id)
            if v:
                po.vendor_name = v.name
                po.vendor_gst  = v.gst or ""
                po.vendor_addr = v.address or ""
                po.vendor_bank = _fmt_bank(v)

        # Replace line items
        if items_data is not None:
            for li in list(po.line_items):
                db.session.delete(li)
            db.session.flush()
            for i, item in enumerate(items_data):
                db.session.add(_build_line_item(po.id, item, i))
            db.session.flush()

        po.recalculate_totals()
        db.session.commit()
        return ok(po.to_dict(), "Purchase Order updated")

    except Exception as e:
        db.session.rollback()
        return server_err(e)


# ─── STATUS TRANSITION ───────────────────────────────────────
@po_bp.patch("/<string:po_id>/status")
def change_status(po_id):
    po = PurchaseOrder.query.get(po_id)
    if not po:
        return not_found("Purchase Order")

    user = current_user()
    if not user:
        return err("Not authenticated — please log in again.", 401)
    role       = user.get("role", "accounts")
    data       = request.get_json(silent=True) or {}
    new_status = data.get("status", "").strip()

    if new_status not in VALID_STATUSES:
        return err(f"Invalid status. Choose from: {', '.join(VALID_STATUSES)}")

    # Role-based workflow:
    # accounts can only move Draft → Pending Approval
    # COO can move Pending Approval → Approved / Rejected, and Approved → Closed
    allowed = {
        "accounts": {
            "Draft":            ["Pending Approval"],
            "Pending Approval": ["Draft"],        # retract back to draft
            "Approved":         ["Closed"],
            "Rejected":         ["Draft"],
            "Closed":           [],
        },
        "coo": {
            "Draft":            ["Pending Approval", "Approved"],
            "Pending Approval": ["Approved", "Rejected", "Draft"],
            "Approved":         ["Closed"],
            "Rejected":         ["Draft"],
            "Closed":           [],
        },
    }

    allowed_transitions = allowed.get(role, allowed["accounts"])
    if new_status not in allowed_transitions.get(po.status, []):
        return err(
            f"Your role ({role}) cannot move a PO from "
            f"'{po.status}' to '{new_status}'.", 403
        )

    try:
        if new_status == "Approved" and data.get("approved_by"):
            po.approved_by = data["approved_by"]
        if new_status == "Rejected":
            po.rejection_reason = data.get("rejection_reason", "").strip()
        if new_status in ("Draft", "Pending Approval"):
            po.rejection_reason = None   # clear on resubmit
        if "coo_remarks" in data:
            po.coo_remarks = (data["coo_remarks"] or "").strip() or None
        po.status = new_status
        db.session.commit()

        if new_status == "Approved":
            from app.mail import send_approval_mail
            send_approval_mail(po)

        return ok({"id": po.id, "status": po.status, "coo_remarks": po.coo_remarks or ""}, f"Status updated to {new_status}")
    except Exception as e:
        db.session.rollback()
        return server_err(e)
    
# ─── REVISE PO ────────────────────────────────────────────────
@po_bp.post("/<string:po_id>/revise")
@login_required
def revise_po(po_id):
    original = PurchaseOrder.query.get(po_id)
    if not original:
        return not_found("Purchase Order")

    if original.status not in ("Approved", "Rejected", "Pending Approval"):
        return err("Only Approved, Rejected or Pending Approval POs can be revised", 409)

    # Check if a revision already exists
    existing = PurchaseOrder.query.filter(
        PurchaseOrder.id.like(f"{po_id}-R%")
    ).order_by(PurchaseOrder.id.desc()).first()

    if existing:
        # Get revision number and increment
        try:
            rev_num = int(existing.id.split("-R")[-1]) + 1
        except ValueError:
            rev_num = 2
    else:
        rev_num = 1

    new_id = f"{po_id}-R{rev_num}"

    try:
        # Close the original
        original.status = "Closed"
        original.notes  = (original.notes or "") + f"\n[Revised → {new_id}]"

        # Create revised copy
        revised = PurchaseOrder(
            id            = new_id,
            vendor_id     = original.vendor_id,
            vendor_name   = original.vendor_name,
            vendor_gst    = original.vendor_gst,
            vendor_addr   = original.vendor_addr,
            vendor_bank   = original.vendor_bank,
            department    = original.department,
            requested_by  = original.requested_by,
            created_by    = original.created_by,
            approved_by   = "",
            po_date       = original.po_date,
            delivery_date = original.delivery_date,
            payment_terms = original.payment_terms,
            notes         = f"[Revised from {po_id}]\n" + (original.notes or "").replace(f"\n[Revised → {new_id}]", ""),
            status        = "Draft",
            advance_pct   = float(original.advance_pct or 0),
            order_type    = original.order_type,
            tds_pct       = float(original.tds_pct or 0),
        )
        db.session.add(revised)
        db.session.flush()

        # Copy line items
        for li in original.line_items:
            new_li = LineItem(
                po_id        = new_id,
                sort_order   = li.sort_order,
                item_name    = li.item_name,
                description  = li.description,
                hsn_code     = li.hsn_code,
                department   = li.department,
                qty          = float(li.qty or 1),
                mrp          = float(li.mrp or 0),
                unit_price   = float(li.unit_price or 0),
                discount_pct = float(li.discount_pct or 0),
                gst_pct      = float(li.gst_pct or 18),
            )
            new_li.compute()
            db.session.add(new_li)

        db.session.flush()
        revised.recalculate_totals()
        db.session.commit()

        return created(revised.to_dict(), f"Revised PO {new_id} created from {po_id}")

    except Exception as e:
        db.session.rollback()
        return server_err(e)

# ─── COO REMARKS ─────────────────────────────────────────────
@po_bp.patch("/<string:po_id>/remarks")
def save_coo_remarks(po_id):
    user = current_user()
    if not user:
        return err("Not authenticated.", 401)
    if user.get("role") not in ("coo", "admin"):
        return err("Only the COO can save remarks.", 403)
    po = PurchaseOrder.query.get(po_id)
    if not po:
        return not_found("Purchase Order")
    data = request.get_json(silent=True) or {}
    po.coo_remarks = (data.get("coo_remarks") or "").strip() or None
    try:
        db.session.commit()
        return ok({"coo_remarks": po.coo_remarks or ""}, "Remarks saved")
    except Exception as e:
        db.session.rollback()
        return server_err(e)


# ─── DELETE ──────────────────────────────────────────────────
@po_bp.delete("/<string:po_id>")
def delete_po(po_id):
    po = PurchaseOrder.query.get(po_id)
    if not po:
        return not_found("Purchase Order")

    if po.status != "Draft":
        return err("Only Draft POs can be deleted", 409)

    try:
        db.session.delete(po)
        db.session.commit()
        return ok(msg=f"PO {po_id} deleted")
    except Exception as e:
        db.session.rollback()
        return server_err(e)


# ─── Private helpers ─────────────────────────────────────────

def _fmt_bank(v: Vendor) -> str:
    parts = [v.bank_name or "", f"A/c {v.bank_acc}" if v.bank_acc else "",
             f"IFSC {v.bank_ifsc}" if v.bank_ifsc else ""]
    return " · ".join(p for p in parts if p)


def _validate_items(items):
    errors = []
    for i, item in enumerate(items):
        if not (item.get("item_name") or "").strip():
            errors.append(f"Row {i+1}: item_name is required")
        try:
            price = float(item.get("unit_price", 0) or 0)
            if price < 0:
                errors.append(f"Row {i+1}: unit_price cannot be negative")
        except (TypeError, ValueError):
            errors.append(f"Row {i+1}: unit_price must be a number")
    return errors


def _build_line_item(po_id, item, index) -> LineItem:
    li = LineItem(
        po_id        = po_id,
        sort_order   = index,
        item_name    = (item.get("item_name")  or "").strip(),
        description  = (item.get("description") or "").strip(),
        hsn_code     = (item.get("hsn_code")    or "").strip(),
        department   = (item.get("department")  or "").strip(),
        qty          = float(item.get("qty",          1)  or 1),
        mrp          = float(item.get("mrp",          0)  or 0),
        unit_price   = float(item.get("unit_price",   0)  or 0),
        discount_pct = float(item.get("discount_pct", 0)  or 0),
        gst_pct      = float(item.get("gst_pct",      18) or 18),
    )
    li.compute()
    return li